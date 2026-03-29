#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 数据库列表读取模块
从 Excel 文件读取数据库连接信息
"""

import logging
from typing import List, Dict, Optional
from dataclasses import dataclass

logger = logging.getLogger(__name__)

@dataclass
class DatabaseInfo:
    """数据库连接信息"""
    platform: str  # "aws", "azure", "other"
    database_type: str  # "mysql" or "postgresql"
    name: str  # 数据库名称/标识
    host: str  # 主机地址
    port: int  # 端口
    username: str  # 主用户名
    password: str  # 主用户密码
    dba_username: str = ""  # 要创建的 DBA 用户名（可选，留空则使用统一配置）
    dba_password: str = ""  # DBA 密码（可选，留空则使用统一配置）
    notes: str = ""  # 备注

class ExcelDatabaseReader:
    """
    Excel 数据库列表读取器
    
    支持的 Excel 格式：
    | platform | database_type | name | host | port | username | password | dba_username | dba_password | notes |
    |----------|---------------|------|------|------|----------|----------|--------------|--------------|-------|
    | aws | mysql | prod-db-1 | rm-xxx.mysql.rds.amazonaws.com | 3306 | admin | password123 | dba_admin | DBAPass123! | 生产数据库 |
    | azure | postgresql | test-db | test.postgres.database.azure.com | 5432 | azureadmin | azpass | | | 测试数据库 |
    """
    
    def __init__(self, excel_path: str):
        """
        初始化 Excel 读取器
        
        Args:
            excel_path: Excel 文件路径
        """
        self.excel_path = excel_path
        self.databases: List[DatabaseInfo] = []
    
    def read_excel(self) -> List[DatabaseInfo]:
        """
        读取 Excel 文件
        
        Returns:
            DatabaseInfo 列表
        """
        logger.info(f"正在读取 Excel 文件：{self.excel_path}")
        
        try:
            import openpyxl
        except ImportError:
            logger.error("需要安装 openpyxl: pip3 install openpyxl")
            raise
        
        try:
            # 加载 Excel 文件
            wb = openpyxl.load_workbook(self.excel_path, read_only=True)
            ws = wb.active
            
            # 读取表头
            headers = []
            for cell in ws[1]:
                headers.append(str(cell.value).strip().lower() if cell.value else "")
            
            logger.info(f"Excel 表头：{headers}")
            
            # 验证必要列
            required_columns = ['platform', 'database_type', 'name', 'host', 'port', 'username', 'password']
            missing_columns = [col for col in required_columns if col not in headers]
            
            if missing_columns:
                raise ValueError(f"Excel 文件缺少必要列：{missing_columns}")
            
            # 读取数据行
            databases = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # 跳过空行
                if not any(row):
                    continue
                
                try:
                    # 创建字典
                    row_dict = dict(zip(headers, row))
                    
                    # 提取字段
                    platform = str(row_dict.get('platform', '')).strip().lower()
                    database_type = str(row_dict.get('database_type', '')).strip().lower()
                    name = str(row_dict.get('name', '')).strip()
                    host = str(row_dict.get('host', '')).strip()
                    port = int(row_dict.get('port', 3306))
                    username = str(row_dict.get('username', '')).strip()
                    password = str(row_dict.get('password', '')).strip()
                    
                    # 可选字段
                    dba_username = str(row_dict.get('dba_username', '')).strip()
                    dba_password = str(row_dict.get('dba_password', '')).strip()
                    notes = str(row_dict.get('notes', '')).strip() if row_dict.get('notes') else ""
                    
                    # 验证必要字段
                    if not platform or not database_type or not name or not host or not username or not password:
                        logger.warning(f"第 {row_idx} 行缺少必要字段，跳过")
                        continue
                    
                    # 验证数据库类型
                    if database_type not in ['mysql', 'postgresql', 'postgres']:
                        logger.warning(f"第 {row_idx} 行数据库类型不支持：{database_type}，跳过")
                        continue
                    
                    # 标准化 postgresql
                    if database_type == 'postgres':
                        database_type = 'postgresql'
                    
                    # 创建 DatabaseInfo 对象
                    db_info = DatabaseInfo(
                        platform=platform,
                        database_type=database_type,
                        name=name,
                        host=host,
                        port=port,
                        username=username,
                        password=password,
                        dba_username=dba_username,
                        dba_password=dba_password,
                        notes=notes
                    )
                    
                    databases.append(db_info)
                    
                except Exception as e:
                    logger.error(f"第 {row_idx} 行解析失败：{str(e)}")
                    continue
            
            wb.close()
            
            self.databases = databases
            logger.info(f"✓ 成功读取 {len(databases)} 个数据库连接")
            
            # 统计信息
            platform_count = {}
            type_count = {}
            for db in databases:
                platform_count[db.platform] = platform_count.get(db.platform, 0) + 1
                type_count[db.database_type] = type_count.get(db.database_type, 0) + 1
            
            logger.info(f"平台分布：{platform_count}")
            logger.info(f"数据库类型分布：{type_count}")
            
        except FileNotFoundError:
            logger.error(f"Excel 文件不存在：{self.excel_path}")
            raise
        except Exception as e:
            logger.error(f"读取 Excel 文件失败：{str(e)}")
            raise
        
        return databases
    
    def get_databases_by_platform(self, platform: str) -> List[DatabaseInfo]:
        """
        按平台筛选数据库
        
        Args:
            platform: 平台名称 ("aws", "azure", "all")
        
        Returns:
            DatabaseInfo 列表
        """
        if platform == "all":
            return self.databases
        else:
            return [db for db in self.databases if db.platform == platform]
    
    def get_databases_by_type(self, database_type: str) -> List[DatabaseInfo]:
        """
        按数据库类型筛选
        
        Args:
            database_type: 数据库类型 ("mysql", "postgresql")
        
        Returns:
            DatabaseInfo 列表
        """
        return [db for db in self.databases if db.database_type == database_type]
    
    def validate_connections(self) -> Dict[str, List[str]]:
        """
        验证数据库连接信息
        
        Returns:
            验证结果字典，包含错误和警告列表
        """
        results = {
            "errors": [],
            "warnings": []
        }
        
        for i, db in enumerate(self.databases, 1):
            # 检查平台
            if db.platform not in ['aws', 'azure', 'other']:
                results["warnings"].append(f"第 {i} 行：未知平台 '{db.platform}'")
            
            # 检查端口范围
            if not (1 <= db.port <= 65535):
                results["errors"].append(f"第 {i} 行：端口号无效 {db.port}")
            
            # 检查主机格式
            if not db.host or '.' not in db.host:
                results["warnings"].append(f"第 {i} 行：主机名格式可能不正确 '{db.host}'")
            
            # 检查密码强度
            if len(db.password) < 8:
                results["warnings"].append(f"第 {i} 行：密码长度过短（{len(db.password)} 位）")
            
            # 检查用户名
            if not db.username:
                results["errors"].append(f"第 {i} 行：用户名为空")
        
        return results
    
    def print_summary(self):
        """打印数据库列表摘要"""
        print("\n" + "=" * 80)
        print("数据库列表摘要")
        print("=" * 80)
        
        print(f"\n总计：{len(self.databases)} 个数据库")
        
        # 按平台统计
        print("\n按平台分布:")
        platform_count = {}
        for db in self.databases:
            platform_count[db.platform] = platform_count.get(db.platform, 0) + 1
        
        for platform, count in platform_count.items():
            print(f"  - {platform}: {count} 个")
        
        # 按类型统计
        print("\n按数据库类型分布:")
        type_count = {}
        for db in self.databases:
            type_count[db.database_type] = type_count.get(db.database_type, 0) + 1
        
        for db_type, count in type_count.items():
            print(f"  - {db_type}: {count} 个")
        
        print("\n" + "=" * 80)
