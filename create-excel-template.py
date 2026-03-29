#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
创建 Excel 数据库列表模板
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def create_template(filename="databases-template.xlsx"):
    """创建 Excel 模板文件"""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Databases"
    
    # 设置列宽
    column_widths = {
        'A': 12,  # platform
        'B': 15,  # database_type
        'C': 20,  # name
        'D': 40,  # host
        'E': 8,   # port
        'F': 15,  # username
        'G': 20,  # password
        'H': 15,  # dba_username
        'I': 20,  # dba_password
        'J': 30,  # notes
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 表头
    headers = [
        'platform',
        'database_type',
        'name',
        'host',
        'port',
        'username',
        'password',
        'dba_username',
        'dba_password',
        'notes'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 示例数据
    examples = [
        {
            'platform': 'aws',
            'database_type': 'mysql',
            'name': 'prod-mysql-01',
            'host': 'rm-prod01.mysql.rds.amazonaws.com',
            'port': 3306,
            'username': 'admin',
            'password': 'YourPassword123!',
            'dba_username': 'dba_admin',
            'dba_password': 'DBAPass123!',
            'notes': '生产 MySQL 数据库'
        },
        {
            'platform': 'aws',
            'database_type': 'postgresql',
            'name': 'prod-pg-01',
            'host': 'pg-prod01.postgres.rds.amazonaws.com',
            'port': 5432,
            'username': 'postgres',
            'password': 'YourPassword456!',
            'dba_username': 'dba_admin',
            'dba_password': 'DBAPass123!',
            'notes': '生产 PostgreSQL 数据库'
        },
        {
            'platform': 'azure',
            'database_type': 'mysql',
            'name': 'test-mysql',
            'host': 'test-mysql.mysql.database.azure.com',
            'port': 3306,
            'username': 'azureadmin',
            'password': 'AzurePass789!',
            'dba_username': '',
            'dba_password': '',
            'notes': '测试 MySQL（使用统一 DBA 凭证）'
        },
        {
            'platform': 'azure',
            'database_type': 'postgresql',
            'name': 'test-pg',
            'host': 'test-pg.postgres.database.azure.com',
            'port': 5432,
            'username': 'azureadmin',
            'password': 'AzurePass012!',
            'dba_username': '',
            'dba_password': '',
            'notes': '测试 PostgreSQL（使用统一 DBA 凭证）'
        },
        {
            'platform': 'other',
            'database_type': 'mysql',
            'name': 'local-db',
            'host': '192.168.1.100',
            'port': 3306,
            'username': 'root',
            'password': 'LocalPass345!',
            'dba_username': 'dba_local',
            'dba_password': 'LocalDBA678!',
            'notes': '本地数据库'
        },
    ]
    
    # 数据行样式
    data_alignment = Alignment(horizontal="left", vertical="center")
    
    for row_idx, data in enumerate(examples, 2):
        for col_idx, key in enumerate(headers, 1):
            value = data.get(key, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = data_alignment
            
            # 偶数行背景色
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
    
    # 保存文件
    wb.save(filename)
    print(f"✅ Excel 模板已创建：{filename}")
    print(f"\n包含 {len(examples)} 个示例数据库")
    print(f"\n使用说明:")
    print(f"  1. 打开 {filename}")
    print(f"  2. 根据实际情况修改示例数据")
    print(f"  3. 删除不需要的示例行")
    print(f"  4. 添加你的数据库信息")
    print(f"  5. 保存并运行：python3 batch-add-dba-users.py")
    print(f"\n必需列：platform, database_type, name, host, port, username, password")
    print(f"可选列：dba_username, dba_password, notes")

if __name__ == "__main__":
    create_template()
